import gc
import logging
import os

import tiktoken
import openpyxl
import openpyxl.reader.excel as _xl_reader

_enc = tiktoken.get_encoding("cl100k_base")
_MAX_TOKENS = 400  # leave headroom for the model's 256 word-piece limit
log = logging.getLogger("parsers.xlsx")


# --- Work around an openpyxl 3.1.x crash on chartsheets ------------------------
# read_chartsheet() does rels.find() on a plain list when a chartsheet has no
# _rels, aborting the ENTIRE load_workbook(). Skip the bad chartsheet so the
# workbook's data sheets still load (stays in low-memory read_only mode).
_orig_read_chartsheet = _xl_reader.ExcelReader.read_chartsheet


def _safe_read_chartsheet(self, sheet, rel):
    try:
        return _orig_read_chartsheet(self, sheet, rel)
    except Exception:
        return


_xl_reader.ExcelReader.read_chartsheet = _safe_read_chartsheet
# ------------------------------------------------------------------------------


def parse_xlsx(file_path: str) -> list[dict]:
    """
    Parse an Excel workbook, iterating all data sheets. The header row is the
    first row that has any non-empty cell (real-world sheets often start with a
    blank or title row before the header -- assuming row 1 is the header made
    those whole files index as empty). Rows are batched under _MAX_TOKENS.

    read_only=True keeps it streaming; chart-only sheets are skipped; memory is
    released via close() + gc.collect() on every exit path. Logs a one-line
    XLSX-EMPTY diagnostic (sheets/kept/skipped) when nothing was extracted.
    """
    wb = None
    chunks: list[dict] = []
    total_sheets = kept = skipped_nondata = 0
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

        for sheet_name in wb.sheetnames:
            total_sheets += 1
            ws = wb[sheet_name]
            if not hasattr(ws, "iter_rows"):     # chart-only / non-data sheet
                skipped_nondata += 1
                continue

            row_iter = ws.iter_rows(values_only=True)

            # Header = first row with any non-empty cell (skip leading blank/title
            # rows). This is the key fix for real files that don't start at A1.
            header_row = None
            for r in row_iter:
                if any(c is not None and str(c).strip() for c in r):
                    header_row = r
                    break
            if header_row is None:
                skipped_nondata += 1
                continue

            headers = [str(h).strip() if h is not None else "" for h in header_row]
            if not any(headers):
                skipped_nondata += 1
                continue

            kept += 1
            header_line = "TABLE: " + " | ".join(h for h in headers if h)
            context_prefix = f"[Context: Sheet — {sheet_name}]\n\n"

            batch_lines: list[str] = []
            batch_tokens = len(_enc.encode(context_prefix + header_line))

            def _flush(lines, _sheet=sheet_name, _prefix=context_prefix, _hdr=header_line):
                if not lines:
                    return
                text = _prefix + _hdr + "\n" + "\n".join(lines)
                chunks.append({"type": "table", "text": text, "sheet_name": _sheet})

            for row in row_iter:
                if all(v is None for v in row):
                    continue

                row_lines = []
                for h, v in zip(headers, row):
                    if h:
                        row_lines.append(f"{h}: {str(v or '').strip()}")
                row_lines.append("---")

                row_tokens = len(_enc.encode("\n".join(row_lines)))

                if batch_lines and batch_tokens + row_tokens > _MAX_TOKENS:
                    _flush(batch_lines)
                    batch_lines = []
                    batch_tokens = len(_enc.encode(context_prefix + header_line))

                batch_lines.extend(row_lines)
                batch_tokens += row_tokens

            _flush(batch_lines)
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass
        gc.collect()

    if not chunks:
        log.info("XLSX-EMPTY %s: sheets=%d kept=%d skipped_nondata=%d",
                 os.path.basename(file_path), total_sheets, kept, skipped_nondata)
    return chunks