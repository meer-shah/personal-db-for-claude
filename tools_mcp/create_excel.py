"""
create_spreadsheet tool — build a real .xlsx from structured content
using openpyxl, upload to OneDrive.
"""

import io

from fastapi import APIRouter, Depends, Request
from pydantic import BaseModel, Field
from slowapi import Limiter
from slowapi.util import get_remote_address

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from tools_mcp.auth import require_bearer
from tools_mcp._onedrive_upload import upload_bytes

router  = APIRouter()
limiter = Limiter(key_func=get_remote_address)

_XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


class ExcelSheet(BaseModel):
    name:    str = "Sheet1"
    headers: list[str] = Field(default_factory=list)
    rows:    list[list] = Field(default_factory=list)  # cells may be str/int/float/bool


class CreateExcelRequest(BaseModel):
    filename:    str
    folder_path: str
    sheets:      list[ExcelSheet] = Field(default_factory=list)


class CreateExcelResponse(BaseModel):
    success:      bool
    onedrive_url: str | None = None
    file_id:      str | None = None
    error:        str | None = None


def _autosize(ws, ncols: int) -> None:
    for col_idx in range(1, ncols + 1):
        letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[letter]:
            v = cell.value
            if v is None:
                continue
            l = len(str(v))
            if l > max_len:
                max_len = l
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 60)


def _build_xlsx(req: CreateExcelRequest) -> bytes:
    wb = Workbook()
    # Remove default sheet — we'll add the user's sheets
    wb.remove(wb.active)

    sheets = req.sheets or [ExcelSheet(name="Sheet1")]
    used_names: set[str] = set()

    for sheet in sheets:
        # Ensure unique, valid sheet name (Excel max 31 chars, no []:*?/\)
        base = sheet.name or "Sheet"
        for ch in '[]:*?/\\':
            base = base.replace(ch, "_")
        base = base[:31] or "Sheet"
        name = base
        i = 2
        while name in used_names:
            suffix = f" ({i})"
            name = (base[: 31 - len(suffix)]) + suffix
            i += 1
        used_names.add(name)

        ws = wb.create_sheet(title=name)

        ncols = max(len(sheet.headers), max((len(r) for r in sheet.rows), default=0))
        if ncols == 0:
            continue

        if sheet.headers:
            ws.append(sheet.headers + [""] * (ncols - len(sheet.headers)))
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill("solid", fgColor="305496")
            header_align = Alignment(horizontal="center", vertical="center")
            for col in range(1, ncols + 1):
                cell = ws.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
            ws.freeze_panes = "A2"

        for row in sheet.rows:
            padded = list(row) + [None] * (ncols - len(row))
            ws.append(padded)

        _autosize(ws, ncols)

    if not wb.sheetnames:
        wb.create_sheet(title="Sheet1")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@router.post("/tools/create_spreadsheet", response_model=CreateExcelResponse)
@limiter.limit("10/minute")
def create_spreadsheet(
    request: Request,
    req: CreateExcelRequest,
    _: None = Depends(require_bearer),
) -> CreateExcelResponse:
    try:
        filename = req.filename if req.filename.lower().endswith(".xlsx") else f"{req.filename}.xlsx"
        raw = _build_xlsx(req)
        data = upload_bytes(raw, filename, req.folder_path, _XLSX_MIME)
        return CreateExcelResponse(
            success      = True,
            onedrive_url = data.get("webUrl"),
            file_id      = data.get("id"),
        )
    except Exception as e:
        return CreateExcelResponse(success=False, error=str(e))
