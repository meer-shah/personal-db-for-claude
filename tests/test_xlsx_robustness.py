import openpyxl
from parsers.xlsx_parser import parse_xlsx
from ingestion import runner


def test_normal_workbook_parses(tmp_path):
    wb = openpyxl.Workbook(); ws = wb.active
    ws["A1"] = "Name"; ws["B1"] = "Val"; ws["A2"] = "x"; ws["B2"] = 1
    f = tmp_path / "n.xlsx"; wb.save(str(f))
    assert len(parse_xlsx(str(f))) >= 1


def test_chartsheet_workbook_does_not_crash(tmp_path):
    # A workbook containing a chartsheet used to crash load_workbook entirely
    # ('list' object has no attribute 'find'). Now the chartsheet is skipped and
    # the data sheet is still extracted.
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Data"
    ws["A1"] = "Name"; ws["B1"] = "Val"; ws["A2"] = "x"; ws["B2"] = 1
    wb.create_chartsheet("ChartOnly")
    f = tmp_path / "chart.xlsx"; wb.save(str(f))
    chunks = parse_xlsx(str(f))
    assert len(chunks) >= 1   # data recovered, no exception


def test_parse_error_quarantines_immediately(tmp_path, monkeypatch):
    monkeypatch.setattr(runner, "QUARANTINE_FILE", tmp_path / "bad.json")
    runner._quarantine_cache = None
    n = runner._record_failure("item1", "f.xlsx", "h1", "PARSE: boom", immediate=True)
    assert n == runner.QUARANTINE_THRESHOLD
    skip, _ = runner._is_quarantined("item1", "h1")
    assert skip is True
    runner._quarantine_cache = None


def test_blank_first_row_is_recovered(tmp_path):
    wb = openpyxl.Workbook(); ws = wb.active
    ws["A2"] = "Name"; ws["B2"] = "Qty"; ws["A3"] = "Bolt"; ws["B3"] = 10  # row 1 blank
    f = tmp_path / "b.xlsx"; wb.save(str(f))
    assert len(parse_xlsx(str(f))) >= 1


def test_title_row_then_header_is_recovered(tmp_path):
    wb = openpyxl.Workbook(); ws = wb.active
    ws["A1"] = "Bill of Materials"            # single-cell title row
    ws["A3"] = "Part"; ws["B3"] = "Cost"; ws["A4"] = "Widget"; ws["B4"] = 99
    f = tmp_path / "t.xlsx"; wb.save(str(f))
    assert len(parse_xlsx(str(f))) >= 1
